#!/usr/bin/env python3
"""
qs2_export_results_excel.py  —  Export QS-QMAODV simulation results to Excel
Chạy trên VM:
    pip install openpyxl --break-system-packages
    python3 ~/qs2_export_results_excel.py

Output: ~/QS-QMAODV-results/v5/QS-QMAODV_Results.xlsx

Sheets:
  1. README          — mô tả file, ngày xuất
  2. Family_N        — raw data + per-condition mean/std
  3. Family_L
  4. Family_E2
  5. Family_C
  6. Family_S
  7. Family_W        — w3 sensitivity (QS2MAODV only)
  8. Family_W2       — ACK-silence threshold sweep
  9. Family_W3       — decay factor sweep
  10. Summary_Stats  — mean ± std mỗi protocol × family
  11. Stat_Tests     — Mann-Whitney U + Cliff's Delta
  12. Ablation       — V1–V5 ablation results (hardcoded từ paper Table 8)
"""

import os, sys
from datetime import datetime
import pandas as pd
import numpy as np

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("openpyxl not found. Installing...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════
DATA_DIR  = os.path.expanduser('~/QS-QMAODV-results/v5')
OUT_FILE  = os.path.join(DATA_DIR, 'QS-QMAODV_Results.xlsx')
PROTOS    = ['AODV', 'PMAODV', 'QMAODV', 'QS2MAODV']

# ── Color palette ─────────────────────────────────────────────────────────────
C_HEADER   = 'FF1A5276'   # dark blue — column headers
C_SUBHEAD  = 'FF2E86C1'   # medium blue — sub-headers
C_QS2      = 'FFFDEDEC'   # light red bg — QS2MAODV rows (proposed)
C_QMAODV   = 'FFEAFAF1'   # light green
C_PMAODV   = 'FFFEF9E7'   # light yellow
C_AODV     = 'FFEAF4FB'   # light blue
C_ALT      = 'FFF2F3F4'   # alternating row
C_TITLE    = 'FF154360'   # sheet title

PROTO_BG = {
    'AODV':    C_AODV,
    'PMAODV':  C_PMAODV,
    'QMAODV':  C_QMAODV,
    'QS2MAODV':C_QS2,
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr_font(bold=True, color='FFFFFFFF', sz=10):
    return Font(bold=bold, color=color, size=sz, name='Calibri')

def cell_font(bold=False, color='FF000000', sz=10):
    return Font(bold=bold, color=color, size=sz, name='Calibri')

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='FFB2BEC3')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def num_fmt(ws, col_letter, row_start, row_end, fmt='0.00'):
    for r in range(row_start, row_end + 1):
        ws[f'{col_letter}{r}'].number_format = fmt

def auto_col_width(ws, min_w=8, max_w=28):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def style_header_row(ws, row, n_cols, bg=C_HEADER):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = hdr_font(bold=True, color='FFFFFFFF')
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = border_thin()

def style_data_row(ws, row, n_cols, proto=None, alt=False):
    bg = PROTO_BG.get(proto, C_ALT if alt else 'FFFFFFFF')
    bold = (proto == 'QS2MAODV')
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = cell_font(bold=bold)
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border_thin()


# ══════════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════
def load(name):
    path = os.path.join(DATA_DIR, f'family_{name}.csv')
    if not os.path.exists(path):
        return None
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    if 'PDR' in df.columns and df['PDR'].max() <= 1.01:
        df['PDR'] = df['PDR'] * 100.0
    return df

families = {
    'N':  load('N'),
    'L':  load('L'),
    'E2': load('E2'),
    'C':  load('C'),
    'S':  load('S'),
    'W':  load('W'),
    'W2': load('W2'),
    'W3': load('W3'),
}
available = {k: v for k, v in families.items() if v is not None}
print(f"Loaded: {list(available.keys())}")


# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)   # xoá sheet mặc định

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 1: README
# ──────────────────────────────────────────────────────────────────────────────
ws_r = wb.create_sheet('README')
ws_r.sheet_view.showGridLines = False
ws_r.column_dimensions['A'].width = 30
ws_r.column_dimensions['B'].width = 60

info = [
    ('Project',       'QS-QMAODV: Queue-State-Aware Q-Learning Multipath AODV'),
    ('Exported',      datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('Author',        'Le Trong Hien  <tronghien1011@gmail.com>'),
    ('Simulator',     'NS-3.40, 30 seeds per condition, 200 s simulation time'),
    ('Protocols',     'AODV  |  PMAODV  |  QMAODV  |  QS2MAODV (proposed)'),
    ('',              ''),
    ('Sheets',        ''),
    ('Family_N',      'Node Density: nNodes ∈ {10,20,30,40,50}'),
    ('Family_L',      'Traffic Load: pktInterval ∈ {1.0,0.5,0.25,0.1,0.05} s'),
    ('Family_E2',     'Energy: InitEnergy ∈ {1,2,5,10,20} J'),
    ('Family_C',      'Combined Stress: 5 load+mobility levels'),
    ('Family_S',      'Node Speed: speed ∈ {5,15,25,35,45} m/s'),
    ('Family_W',      'w3 Sensitivity: w3 ∈ {0.00…0.50}, QS2MAODV only'),
    ('Family_W2',     'ACK-Silence Threshold: {5,10,15,20,30} s'),
    ('Family_W3',     'Decay Factor: {0.85,0.90,0.92,0.95,0.99}'),
    ('Summary_Stats', 'Mean ± Std per Protocol × Family'),
    ('Stat_Tests',    'Mann-Whitney U (one-sided) + Cliff\'s Delta effect size'),
    ('Ablation',      'V1–V5 ablation variants (from paper Table 8)'),
    ('',              ''),
    ('PDR note',      'All PDR values shown as percentage (0–100 %)'),
    ('NRL note',      'Normalized Routing Load = routing_pkts / data_pkts'),
    ('Effect size',   'Cliff\'s δ: |δ|<0.147 negligible, <0.33 small, <0.474 medium, ≥0.474 large'),
]

ws_r['A1'] = 'QS-QMAODV Simulation Results'
ws_r['A1'].font = Font(bold=True, size=14, color=C_TITLE, name='Calibri')
ws_r.row_dimensions[1].height = 22

for i, (k, v) in enumerate(info, start=3):
    ws_r.cell(row=i, column=1, value=k).font = cell_font(bold=bool(k), sz=10)
    ws_r.cell(row=i, column=2, value=v).font = cell_font(sz=10)

ws_r.freeze_panes = None


# ──────────────────────────────────────────────────────────────────────────────
# FAMILY SHEETS  (raw data + per-condition aggregate)
# ──────────────────────────────────────────────────────────────────────────────
FAMILY_META = {
    'N':  ('Family N — Node Density',    'nNodes',        'Number of Nodes'),
    'L':  ('Family L — Traffic Load',    'PktInterval_s', 'Packet Interval (s)'),
    'E2': ('Family E2 — Energy',         'InitEnergy_J',  'Initial Energy (J)'),
    'C':  ('Family C — Combined Stress', 'PktInterval_s', 'Stress Level'),
    'S':  ('Family S — Node Speed',      'Speed_ms',      'Node Speed (m/s)'),
    'W':  ('Family W — w3 Sensitivity',  'w3',            'w3 weight'),
    'W2': ('Family W2 — ACK Threshold',  'SweepParam',    'ACK-Silence Threshold (s)'),
    'W3': ('Family W3 — Decay Factor',   'SweepParam',    'Decay Factor'),
}

METRICS_SHOW = {
    'PDR':             'PDR (%)',
    'Delay_ms':        'E2E Delay (ms)',
    'Throughput_kbps': 'Throughput (kbps)',
    'NRL':             'NRL',
    'EnergyPerPkt_J':  'Energy/Pkt (J)',
}

def write_family_sheet(wb, key, df):
    title, x_col, x_label = FAMILY_META.get(key, (key, df.columns[0], df.columns[0]))

    # --- handle missing x_col ---
    if x_col not in df.columns:
        # try common aliases
        for alt in ['pktInterval', 'nNodes', 'InitEnergy', 'Speed_ms', 'speed',
                    'SweepParam', 'w3', df.columns[0]]:
            if alt in df.columns:
                x_col = alt
                break

    metrics = [m for m in METRICS_SHOW if m in df.columns]

    ws = wb.create_sheet(f'Family_{key}')
    ws.freeze_panes = 'A3'

    # ── Title row ──
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + len(metrics) * 2)
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFFFF', name='Calibri')
    ws['A1'].fill = fill(C_TITLE)
    ws['A1'].alignment = center()

    # ── PART A: Per-condition aggregate ──────────────────────────────────────
    hdr_cols = [x_label, 'Protocol'] + \
               [f'{METRICS_SHOW[m]} mean' for m in metrics] + \
               [f'{METRICS_SHOW[m]} std'  for m in metrics]
    n_cols = len(hdr_cols)

    # Header
    for c, h in enumerate(hdr_cols, 1):
        cell = ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, n_cols)

    row = 3
    x_vals = sorted(df[x_col].unique()) if x_col in df.columns else [None]
    proto_col = df['Protocol'] if 'Protocol' in df.columns else pd.Series(['QS2MAODV'] * len(df))

    for xv in x_vals:
        sub_x = df[df[x_col] == xv] if xv is not None else df
        protos_in = [p for p in PROTOS if p in sub_x['Protocol'].unique()] \
                    if 'Protocol' in sub_x.columns else ['QS2MAODV']

        for proto in protos_in:
            sub_p = sub_x[sub_x['Protocol'] == proto] if 'Protocol' in sub_x.columns else sub_x
            ws.cell(row=row, column=1, value=xv)
            ws.cell(row=row, column=2, value=proto)
            c = 3
            for m in metrics:
                ws.cell(row=row, column=c,     value=round(sub_p[m].mean(), 3))
                ws.cell(row=row, column=c + len(metrics), value=round(sub_p[m].std(), 3))
                c += 1
            style_data_row(ws, row, n_cols, proto=proto)
            row += 1

    # ── Gap ──
    row += 1

    # ── PART B: Raw data ─────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value='── Raw Data ──')
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True, color=C_TITLE, name='Calibri')
    row += 1

    raw_hdr = list(df.columns)
    for c, h in enumerate(raw_hdr, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(raw_hdr), bg=C_SUBHEAD)
    row += 1

    for ridx, (_, drow) in enumerate(df.iterrows()):
        proto = drow.get('Protocol', 'QS2MAODV')
        for c, v in enumerate(drow.values, 1):
            ws.cell(row=row, column=c, value=round(float(v), 4) if isinstance(v, (float, np.floating)) else v)
        style_data_row(ws, row, len(raw_hdr), proto=proto, alt=(ridx % 2 == 0))
        row += 1

    auto_col_width(ws)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 28
    print(f'  Sheet Family_{key}: {row-1} rows')


for key, df in available.items():
    write_family_sheet(wb, key, df)


# ──────────────────────────────────────────────────────────────────────────────
# SUMMARY STATS SHEET
# ──────────────────────────────────────────────────────────────────────────────
ws_s = wb.create_sheet('Summary_Stats')
ws_s.freeze_panes = 'A3'

metrics_sum = ['PDR', 'Delay_ms', 'Throughput_kbps', 'NRL']

# Build aggregate table: Family × Protocol → mean of each metric
rows_sum = []
for key, df in available.items():
    if 'Protocol' not in df.columns:
        continue
    for proto in PROTOS:
        sub = df[df['Protocol'] == proto]
        if len(sub) == 0:
            continue
        r = {'Family': key, 'Protocol': proto}
        for m in metrics_sum:
            if m in df.columns:
                r[f'{m}_mean'] = round(sub[m].mean(), 3)
                r[f'{m}_std']  = round(sub[m].std(),  3)
        rows_sum.append(r)

df_sum = pd.DataFrame(rows_sum)

# Title
ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_sum.columns))
ws_s['A1'] = 'Summary Statistics — Mean ± Std per Protocol × Family'
ws_s['A1'].font = Font(bold=True, size=12, color='FFFFFFFF', name='Calibri')
ws_s['A1'].fill = fill(C_TITLE)
ws_s['A1'].alignment = center()

if not df_sum.empty:
    hdr = list(df_sum.columns)
    for c, h in enumerate(hdr, 1):
        ws_s.cell(row=2, column=c, value=h)
    style_header_row(ws_s, 2, len(hdr))

    for ridx, (_, row_data) in enumerate(df_sum.iterrows()):
        proto = row_data.get('Protocol', '')
        for c, v in enumerate(row_data.values, 1):
            ws_s.cell(row=ridx + 3, column=c, value=v)
        style_data_row(ws_s, ridx + 3, len(hdr), proto=proto)

auto_col_width(ws_s)
print(f'  Sheet Summary_Stats: {len(df_sum)} rows')


# ──────────────────────────────────────────────────────────────────────────────
# STAT TESTS SHEET
# ──────────────────────────────────────────────────────────────────────────────
try:
    from scipy import stats as scipy_stats

    def cliffs_delta(a, b):
        a, b = np.array(a), np.array(b)
        n = len(a) * len(b)
        if n == 0: return 0.0
        greater = sum(1 for ai in a for bi in b if ai > bi)
        less    = sum(1 for ai in a for bi in b if ai < bi)
        return (greater - less) / n

    def effect_mag(cd):
        a = abs(cd)
        if a < 0.147: return 'negligible'
        if a < 0.330: return 'small'
        if a < 0.474: return 'medium'
        return 'large'

    def sig_label(p):
        if p < 0.001: return '***'
        if p < 0.01:  return '**'
        if p < 0.05:  return '*'
        if p < 0.10:  return '†'
        return 'ns'

    ws_t = wb.create_sheet('Stat_Tests')
    ws_t.freeze_panes = 'A3'

    hdr_t = ['Family', 'Condition', 'Metric',
             'QS2MAODV mean', 'QMAODV mean', 'Δ mean',
             'U statistic', 'p-value', 'Significance',
             "Cliff's δ", 'Effect Size']

    ws_t.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdr_t))
    ws_t['A1'] = 'Statistical Tests — Mann-Whitney U (one-sided, QS2MAODV > QMAODV) + Cliff\'s Delta'
    ws_t['A1'].font = Font(bold=True, size=12, color='FFFFFFFF', name='Calibri')
    ws_t['A1'].fill = fill(C_TITLE)
    ws_t['A1'].alignment = center()

    for c, h in enumerate(hdr_t, 1):
        ws_t.cell(row=2, column=c, value=h)
    style_header_row(ws_t, 2, len(hdr_t))

    stat_rows = []
    for key, df in available.items():
        if 'Protocol' not in df.columns:
            continue
        x_col = FAMILY_META.get(key, (None, df.columns[0]))[1]
        if x_col not in df.columns:
            x_col = None

        conditions = [(None, 'all')] if x_col is None \
                     else [(v, f'{x_col}={v}') for v in sorted(df[x_col].unique())]

        for xv, cond_label in conditions:
            sub = df[df[x_col] == xv] if xv is not None and x_col else df
            for metric in ['PDR', 'NRL']:
                if metric not in df.columns:
                    continue
                a = sub[sub['Protocol'] == 'QS2MAODV'][metric].dropna().values
                b = sub[sub['Protocol'] == 'QMAODV'][metric].dropna().values
                if len(a) < 2 or len(b) < 2:
                    continue
                u, p = scipy_stats.mannwhitneyu(a, b, alternative='greater')
                cd   = cliffs_delta(a, b)
                stat_rows.append({
                    'Family':      key,
                    'Condition':   cond_label,
                    'Metric':      metric,
                    'QS2_mean':    round(a.mean(), 3),
                    'QM_mean':     round(b.mean(), 3),
                    'Delta':       round(a.mean() - b.mean(), 3),
                    'U':           round(u, 1),
                    'p':           round(p, 4),
                    'sig':         sig_label(p),
                    'CliffD':      round(cd, 3),
                    'Effect':      effect_mag(cd),
                })

    for ridx, r in enumerate(stat_rows):
        row_vals = [r['Family'], r['Condition'], r['Metric'],
                    r['QS2_mean'], r['QM_mean'], r['Delta'],
                    r['U'], r['p'], r['sig'], r['CliffD'], r['Effect']]
        for c, v in enumerate(row_vals, 1):
            ws_t.cell(row=ridx + 3, column=c, value=v)
        # Colour sig column
        sig_cell = ws_t.cell(row=ridx + 3, column=9)
        if r['sig'] in ('***', '**', '*'):
            sig_cell.fill = fill('FFD5F5E3')
            sig_cell.font = Font(bold=True, color='FF1D8348', name='Calibri')
        alt = ridx % 2 == 0
        for c in range(1, len(hdr_t) + 1):
            cell = ws_t.cell(row=ridx + 3, column=c)
            if c != 9:  # don't override sig cell fill
                cell.fill = fill(C_ALT if alt else 'FFFFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin()

    auto_col_width(ws_t)
    print(f'  Sheet Stat_Tests: {len(stat_rows)} rows')

except ImportError:
    ws_t = wb.create_sheet('Stat_Tests')
    ws_t['A1'] = 'scipy not installed — run: pip install scipy --break-system-packages'
    print('  [SKIP] Stat_Tests — scipy not available')


# ──────────────────────────────────────────────────────────────────────────────
# ABLATION SHEET
# ──────────────────────────────────────────────────────────────────────────────
ws_a = wb.create_sheet('Ablation')
ws_a.freeze_panes = 'A3'

abl_data = {
    'Variant':           ['V1 (base QMAODV)', 'V2 (+queue-aware)',
                          'V3 (+ACK decay)',  'V4 (+adaptive w)',
                          'V5 (+trend ε)  ← proposed'],
    'PDR (%)':           [30.16, 32.30, 36.37, 36.60, 37.03],
    'E2E Delay (ms)':    [224.1, 215.5, 184.1, 178.2, 173.8],
    'Throughput (kbps)': [128.4, 137.6, 155.1, 156.1, 157.9],
    'NRL':               [82.1,  126.4, 92.0,  88.2,  84.6],
}
df_abl = pd.DataFrame(abl_data)

ws_a.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_abl.columns))
ws_a['A1'] = 'Ablation Study — V1 to V5 Incremental Feature Addition'
ws_a['A1'].font = Font(bold=True, size=12, color='FFFFFFFF', name='Calibri')
ws_a['A1'].fill = fill(C_TITLE)
ws_a['A1'].alignment = center()

abl_hdr = list(df_abl.columns)
for c, h in enumerate(abl_hdr, 1):
    ws_a.cell(row=2, column=c, value=h)
style_header_row(ws_a, 2, len(abl_hdr))

for ridx, (_, row_data) in enumerate(df_abl.iterrows()):
    is_proposed = ridx == len(df_abl) - 1
    for c, v in enumerate(row_data.values, 1):
        cell = ws_a.cell(row=ridx + 3, column=c, value=v)
        cell.fill      = fill(C_QS2 if is_proposed else (C_ALT if ridx % 2 == 0 else 'FFFFFFFF'))
        cell.font      = cell_font(bold=is_proposed)
        cell.alignment = center()
        cell.border    = border_thin()

auto_col_width(ws_a)

# ── Add bar chart for PDR ─────────────────────────────────────────────────────
chart_pdr = BarChart()
chart_pdr.type   = 'col'
chart_pdr.title  = 'Ablation — PDR (%)'
chart_pdr.y_axis.title = 'PDR (%)'
chart_pdr.x_axis.title = 'Variant'
chart_pdr.width  = 15
chart_pdr.height = 10
chart_pdr.style  = 10

data_ref   = Reference(ws_a, min_col=2, max_col=2, min_row=2, max_row=7)
cats_ref   = Reference(ws_a, min_col=1, min_row=3, max_row=7)
chart_pdr.add_data(data_ref, titles_from_data=True)
chart_pdr.set_categories(cats_ref)
ws_a.add_chart(chart_pdr, 'G2')

print(f'  Sheet Ablation: {len(df_abl)} rows + bar chart')


# ──────────────────────────────────────────────────────────────────────────────
# SAVE
# ──────────────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f'\n✓ Excel saved: {OUT_FILE}')
print(f'  Sheets: {[s.title for s in wb.worksheets]}')
